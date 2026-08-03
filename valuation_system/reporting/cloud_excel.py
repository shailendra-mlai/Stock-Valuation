from __future__ import annotations

import json
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NAVY = "17365D"
BLUE = "1F4E78"
WHITE = "FFFFFF"
GRAY = "E7E6E6"
FORECAST = "D9EAF7"
YELLOW = "FFF2CC"
RED = "FCE4D6"
GREEN = "E2F0D9"
PURPLE = "E4DFEC"
GRID = "D9E2F3"
INPUT_BLUE = "0000FF"
LINK_GREEN = "008000"

MONEY = '$#,##0;[Red]($#,##0);-'
PER_SHARE = '$0.00;[Red]($0.00);-'
PERCENT = '0.0%;[Red](0.0%);-'
MULTIPLE = '0.0x;[Red](0.0x);-'
COUNT = '#,##0.0;[Red](#,##0.0);-'

TABS = [
    "Valuation",
    "Income Statement",
    "Balance Sheet",
    "Cash Flow Statement",
    "ROIC & Economic Profit",
    "Disaggregated ROIC",
    "Comparables",
    "ROIC Tree - Detail",
    "ROIC Tree - Peers",
    "Value Drivers & Indicators",
    "Forecast Inputs",
    "Revenue Build",
    "PP&E Schedule",
    "Cost Build",
    "IS Forecast",
    "BS Forecast",
    "DCF",
    "CF Forecast",
    "ROIC Tree - Forecast",
    "APV Method",
    "APV Tax Shields",
    "Financing Plan",
    "Scenarios",
    "Model Checks",
    "Sources",
]


def _excel_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool, date, datetime)):
        return value
    return json.dumps(value, default=str, sort_keys=True)


def _title(ws: Any, title: str, subtitle: str, end_column: int = 8) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_column)
    ws.cell(1, 1, title)
    ws.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    ws.cell(1, 1).font = Font(color=WHITE, bold=True, size=16)
    ws.cell(1, 1).alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_column)
    ws.cell(2, 1, subtitle)
    ws.cell(2, 1).fill = PatternFill("solid", fgColor="DCE6F1")
    ws.cell(2, 1).font = Font(color="333333", italic=True)
    ws.cell(2, 1).alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30


def _header(ws: Any, row: int, start_column: int, end_column: int) -> None:
    thin = Side(style="thin", color=GRID)
    for cell in ws.iter_cols(
        min_col=start_column, max_col=end_column, min_row=row, max_row=row
    ):
        target = cell[0]
        target.fill = PatternFill("solid", fgColor=BLUE)
        target.font = Font(color=WHITE, bold=True)
        target.alignment = Alignment(vertical="center", wrap_text=True)
        target.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 34


def _body(ws: Any, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="E7E6E6")
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.font = Font(color="222222", size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=thin)


def _table(
    ws: Any,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    start_row: int = 4,
    start_column: int = 1,
    freeze: str | None = "A5",
) -> tuple[int, int]:
    rows = list(rows)
    for offset, header in enumerate(headers):
        ws.cell(start_row, start_column + offset, header)
    _header(ws, start_row, start_column, start_column + len(headers) - 1)
    for row_offset, values in enumerate(rows, start=1):
        for column_offset, value in enumerate(values):
            ws.cell(
                start_row + row_offset,
                start_column + column_offset,
                _excel_value(value),
            )
    if rows:
        _body(
            ws,
            start_row + 1,
            start_row + len(rows),
            start_column,
            start_column + len(headers) - 1,
        )
    if freeze:
        ws.freeze_panes = freeze
    return start_row + 1, start_row + len(rows)


def _set_widths(ws: Any, widths: dict[int, float]) -> None:
    for column, width in widths.items():
        ws.column_dimensions[get_column_letter(column)].width = width


def _format(
    ws: Any, min_row: int, max_row: int, min_col: int, max_col: int, number_format: str
) -> None:
    if max_row < min_row:
        return
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.number_format = number_format
            cell.alignment = Alignment(horizontal="right", vertical="center")


def _linked(cell: Any) -> None:
    cell.font = Font(color=LINK_GREEN, bold=True)


def _input(cell: Any) -> None:
    cell.fill = PatternFill("solid", fgColor=YELLOW)
    cell.font = Font(color=INPUT_BLUE)


def _status_rules(ws: Any, cell_range: str) -> None:
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("PASS",{cell_range.split(":")[0]}))'], fill=PatternFill("solid", fgColor=GREEN)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("WARNING",{cell_range.split(":")[0]}))'], fill=PatternFill("solid", fgColor=YELLOW)),
    )
    ws.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=[f'ISNUMBER(SEARCH("FAIL",{cell_range.split(":")[0]}))'], fill=PatternFill("solid", fgColor=RED)),
    )


def _timeseries(
    ws: Any,
    title: str,
    subtitle: str,
    periods: Sequence[dict[str, Any]],
    metrics: Sequence[tuple[str, str, str]],
    *,
    fill: str,
) -> None:
    end_column = max(2, len(periods) + 1)
    _title(ws, title, subtitle, end_column)
    headers = ["Metric", *[row.get("year") for row in periods]]
    rows = [[label, *[row.get(key) for row in periods]] for label, key, _ in metrics]
    first, last = _table(ws, headers, rows)
    for index, (_, _, number_format) in enumerate(metrics, start=first):
        _format(ws, index, index, 2, end_column, number_format)
    for row in ws.iter_rows(min_row=first, max_row=last, min_col=2, max_col=end_column):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=fill)
    _set_widths(ws, {1: 34, **{column: 14 for column in range(2, end_column + 1)}})
    ws.freeze_panes = "B5"


def _latest_company_history(result: Any) -> dict[str, Any]:
    rows = result.company.get("historical", [])
    return rows[-1] if rows else {}


def export_cloud_excel(result: Any, output_path: str | Path) -> Path:
    """Create a Streamlit-compatible, reference-style valuation workbook."""
    path = Path(output_path).resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in TABS:
        workbook.create_sheet(name)

    s = result.summary
    h = result.historical
    f = result.forecast
    op = result.overperformance or []
    a = result.assumptions
    company = result.company
    latest = _latest_company_history(result)
    ticker = result.ticker
    company_name = company.get("name", ticker)
    currency = company.get("currency", a.get("currency", "USD"))
    valuation_date = a.get("valuation_date", "")

    # Valuation summary
    ws = workbook["Valuation"]
    _title(
        ws,
        f"{ticker} Intrinsic Valuation",
        f"{company_name} | APV framework | Valuation date {valuation_date} | {currency} millions",
        10,
    )
    summary_labels = [
        "Unlevered TOCC",
        "Terminal growth",
        "PV explicit FCF ($mm)",
        "PV continuing value ($mm)",
        "Operating enterprise value ($mm)",
        "PV financing effects ($mm)",
        "APV enterprise value ($mm)",
        "Equity value ($mm)",
        "Intrinsic value / share",
        "Market price",
        "Market capitalization ($mm)",
        "Premium / (discount)",
        "PV continuing value / APV EV",
        "Overall model status",
    ]
    _table(ws, ["Valuation summary", "Value"], [[label, None] for label in summary_labels])
    formulas = [
        "='APV Method'!B5",
        "='APV Method'!B6",
        "='APV Method'!B8",
        "='APV Method'!B9",
        "='APV Method'!B10",
        "='APV Method'!B14",
        "='APV Method'!B15",
        "='APV Method'!B21",
        "='APV Method'!B23",
        "='APV Method'!B24",
        "='APV Method'!B25",
        "='APV Method'!B26",
        "='APV Method'!B16",
        "='Model Checks'!B4",
    ]
    for row, formula in enumerate(formulas, start=5):
        ws.cell(row, 2, formula)
        _linked(ws.cell(row, 2))
    _format(ws, 5, 6, 2, 2, PERCENT)
    _format(ws, 7, 12, 2, 2, MONEY)
    _format(ws, 13, 14, 2, 2, PER_SHARE)
    _format(ws, 15, 15, 2, 2, MONEY)
    _format(ws, 16, 17, 2, 2, PERCENT)
    _status_rules(ws, "B18")
    ws.merge_cells("D4:J4")
    ws["D4"] = "Model conventions and workbook navigation"
    _header(ws, 4, 4, 10)
    conventions = [
        "Operating value is discounted at unlevered TOCC.",
        "Explicit, competitive-advantage fade, and steady-state periods are separated.",
        "Terminal RONIC converges to TOCC; no perpetual excess return is assumed.",
        "Tax shields are valued separately using parallel NOL schedules.",
        "Liquidity shortfalls are disclosed rather than silently funded.",
        "Blue-font yellow cells are editable assumptions or probabilities.",
        "Green-font values link to another worksheet.",
        "Review Sources, Value Drivers, Scenarios, and Model Checks before use.",
    ]
    for row, text in enumerate(conventions, start=5):
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=10)
        ws.cell(row, 4, text)
        ws.cell(row, 4).fill = PatternFill("solid", fgColor="F2F2F2")
        ws.cell(row, 4).alignment = Alignment(wrap_text=True, vertical="center")
        ws.row_dimensions[row].height = 28
    _set_widths(ws, {1: 42, 2: 22, **{column: 15 for column in range(4, 11)}})

    # Historical statements
    raw = company.get("historical", [])
    cash_flow_history = []
    for index, row in enumerate(h):
        previous_owc = h[index - 1].get("owc") if index else None
        current_owc = row.get("owc")
        cash_flow_history.append({
            **row,
            "change_owc": (
                current_owc - previous_owc
                if current_owc is not None and previous_owc is not None
                else None
            ),
        })
    _timeseries(
        workbook["Income Statement"],
        "Historical Income Statement",
        "Normalized annual reported values; currency millions",
        h,
        [
            ("Revenue", "revenue", MONEY),
            ("Cost of goods sold", "cogs", MONEY),
            ("Gross profit", "gross_profit", MONEY),
            ("SG&A", "sga", MONEY),
            ("R&D", "rd", MONEY),
            ("D&A", "da", MONEY),
            ("EBIT", "ebit", MONEY),
            ("Operating taxes", "taxes", MONEY),
        ],
        fill=GRAY,
    )
    _timeseries(
        workbook["Balance Sheet"],
        "Historical Balance Sheet",
        "Operating assets, financing claims, and equity; currency millions",
        raw,
        [
            ("Cash", "cash", MONEY),
            ("Marketable securities", "marketable_securities", MONEY),
            ("Receivables", "receivables", MONEY),
            ("Inventory", "inventory", MONEY),
            ("Other current operating assets", "other_current_operating_assets", MONEY),
            ("Net PP&E", "net_ppe", MONEY),
            ("Operating lease assets", "operating_lease_assets", MONEY),
            ("Other operating assets", "other_operating_assets", MONEY),
            ("Accounts payable", "accounts_payable", MONEY),
            ("Accrued operating liabilities", "accrued_operating_liabilities", MONEY),
            ("Deferred revenue", "deferred_revenue", MONEY),
            ("Other operating liabilities", "other_operating_liabilities", MONEY),
            ("Debt", "debt", MONEY),
            ("Lease liabilities", "lease_liabilities", MONEY),
            ("Equity", "equity", MONEY),
        ],
        fill=GRAY,
    )
    _timeseries(
        workbook["Cash Flow Statement"],
        "Historical Cash Flow Statement",
        "Operating and investing cash-flow indicators; currency millions",
        cash_flow_history,
        [
            ("NOPAT", "nopat", MONEY),
            ("D&A", "da", MONEY),
            ("Capital expenditures", "capex", MONEY),
            ("Change in operating working capital", "change_owc", MONEY),
            ("Unlevered free cash flow", "fcf", MONEY),
            ("Alternative FCF reconciliation", "fcf_alt", MONEY),
            ("Stock-based compensation", "stock_comp", MONEY),
        ],
        fill=GRAY,
    )

    # Historical ROIC analysis
    ws = workbook["ROIC & Economic Profit"]
    _title(ws, "ROIC and Economic Profit", "Historical value creation versus unlevered TOCC", 10)
    roic_rows = [
        [
            row.get("year"), row.get("revenue"), row.get("nopat"),
            row.get("operating_invested_capital"), row.get("roic"), row.get("tocc"),
            row.get("roic_spread"), row.get("economic_profit"), row.get("fcf"),
            "Created value" if (row.get("roic_spread") or 0) > 0 else "Destroyed value",
        ]
        for row in h
    ]
    first, last = _table(
        ws,
        ["Year", "Revenue", "NOPAT", "Operating invested capital", "ROIC", "TOCC", "ROIC − TOCC", "Economic profit", "UFCF", "Assessment"],
        roic_rows,
    )
    _format(ws, first, last, 2, 4, MONEY)
    _format(ws, first, last, 5, 7, PERCENT)
    _format(ws, first, last, 8, 9, MONEY)
    _set_widths(ws, {1: 12, 2: 18, 3: 18, 4: 24, 5: 15, 6: 15, 7: 18, 8: 20, 9: 18, 10: 20})

    ws = workbook["Disaggregated ROIC"]
    _title(ws, "Disaggregated ROIC", "ROIC = NOPAT margin × operating capital turnover", 10)
    rows = [
        [
            row.get("year"), row.get("gross_margin"), row.get("ebit_margin"),
            row.get("tax_efficiency"), row.get("nopat_margin"), row.get("receivable_days"),
            row.get("inventory_days"), row.get("payable_days"), row.get("capital_turnover"), row.get("roic"),
        ]
        for row in h
    ]
    first, last = _table(ws, ["Year", "Gross margin", "EBIT margin", "Tax efficiency", "NOPAT margin", "Receivable days", "Inventory days", "Payable days", "Capital turnover", "ROIC"], rows)
    _format(ws, first, last, 2, 5, PERCENT)
    _format(ws, first, last, 6, 8, "0.0")
    _format(ws, first, last, 9, 9, MULTIPLE)
    _format(ws, first, last, 10, 10, PERCENT)
    _set_widths(ws, {column: 18 for column in range(1, 11)})

    # Comparable-company analysis
    ws = workbook["Comparables"]
    _title(ws, "Comparable Companies", "Selected peer set used for unlevered operating risk and ROIC comparison", 12)
    peer_headers = ["Peer", "Company", "Equity beta", "Market cap", "Debt", "Raw asset beta", "Adjusted asset beta", "Recommendation score", "Weight", "Revenue", "ROIC", "Source"]
    peer_rows = []
    for row in result.tocc_peers:
        peer_rows.append([
            row.get("peer"), row.get("company_name"), row.get("equity_beta"), row.get("equity"), row.get("debt"),
            row.get("raw_asset_beta"), row.get("adjusted_asset_beta"), row.get("recommendation_score"), row.get("weight"),
            row.get("revenue"), row.get("roic"), row.get("source"),
        ])
    first, last = _table(ws, peer_headers, peer_rows)
    _format(ws, first, last, 3, 3, "0.00x")
    _format(ws, first, last, 4, 5, MONEY)
    _format(ws, first, last, 6, 7, "0.00x")
    _format(ws, first, last, 8, 9, "0.00")
    _format(ws, first, last, 10, 10, MONEY)
    _format(ws, first, last, 11, 11, PERCENT)
    _set_widths(ws, {1: 13, 2: 28, **{column: 18 for column in range(3, 12)}, 12: 48})

    ws = workbook["ROIC Tree - Detail"]
    _title(ws, "ROIC Tree - Detail", "Company value-driver decomposition based on the course ROIC framework", 10)
    detail_rows = [
        [row.get("year"), row.get("roic"), row.get("nopat_margin"), row.get("capital_turnover"), row.get("ebit_margin"), row.get("tax_efficiency"), row.get("gross_margin"), row.get("receivable_days"), row.get("inventory_days"), row.get("payable_days")]
        for row in h
    ]
    first, last = _table(ws, ["Year", "ROIC", "NOPAT margin", "Capital turnover", "EBIT margin", "Tax efficiency", "Gross margin", "Receivable days", "Inventory days", "Payable days"], detail_rows)
    _format(ws, first, last, 2, 3, PERCENT)
    _format(ws, first, last, 4, 4, MULTIPLE)
    _format(ws, first, last, 5, 7, PERCENT)
    _format(ws, first, last, 8, 10, "0.0")
    _set_widths(ws, {column: 18 for column in range(1, 11)})

    ws = workbook["ROIC Tree - Peers"]
    _title(ws, "ROIC Tree - Peers", "Target company versus selected comparable companies", 8)
    target = h[-1] if h else {}
    rows = [[ticker, target.get("roic"), target.get("nopat_margin"), target.get("capital_turnover"), target.get("ebit_margin"), target.get("gross_margin"), target.get("receivable_days"), target.get("inventory_days")]]
    rows.extend([[row.get("peer"), row.get("roic"), row.get("nopat_margin"), row.get("capital_turnover"), row.get("ebit_margin"), row.get("gross_margin"), row.get("receivable_days"), row.get("inventory_days")] for row in result.tocc_peers])
    first, last = _table(ws, ["Company", "ROIC", "NOPAT margin", "Capital turnover", "EBIT margin", "Gross margin", "Receivable days", "Inventory days"], rows)
    _format(ws, first, last, 2, 3, PERCENT)
    _format(ws, first, last, 4, 4, MULTIPLE)
    _format(ws, first, last, 5, 6, PERCENT)
    _format(ws, first, last, 7, 8, "0.0")
    _set_widths(ws, {1: 20, **{column: 18 for column in range(2, 9)}})

    ws = workbook["Value Drivers & Indicators"]
    _title(ws, "Value Drivers and Indicators", "Every forecast input is a falsifiable hypothesis", 9)
    driver_rows = [[row.get("variable"), row.get("historical_evidence"), row.get("management_evidence"), row.get("industry_evidence"), row.get("comparable_evidence"), row.get("base"), row.get("downside"), row.get("upside"), row.get("falsifier")] for row in result.value_drivers]
    first, last = _table(ws, ["Variable", "Historical evidence", "Management evidence", "Industry evidence", "Comparable evidence", "Base", "Downside", "Upside", "What would falsify it?"], driver_rows)
    for row in ws.iter_rows(min_row=first, max_row=last, min_col=1, max_col=9):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _set_widths(ws, {1: 24, 2: 30, 3: 30, 4: 30, 5: 30, 6: 15, 7: 15, 8: 15, 9: 46})

    # Forecast assumptions and operating schedules
    ws = workbook["Forecast Inputs"]
    _title(ws, "Forecast Inputs", "Editable blue-font assumptions; values are hypotheses rather than facts", 4)
    assumption_rows = [
        ["Explicit forecast years", a.get("forecast_years"), "years", "Detailed operating period"],
        ["Competitive-advantage years", a.get("competitive_advantage_years"), "years", "RONIC fade period"],
        ["Starting revenue growth", a.get("revenue_growth_start"), "%", "Historical calibration"],
        ["Terminal revenue growth", a.get("revenue_growth_terminal"), "%", "Normalized operating growth"],
        ["Starting EBIT margin", a.get("ebit_margin_start"), "%", "Historical calibration"],
        ["Terminal EBIT margin", a.get("ebit_margin_terminal"), "%", "Sustainable operating hypothesis"],
        ["Normalized tax rate", a.get("tax_rate"), "%", "Operating cash-tax rate"],
        ["Initial operating NOL", a.get("initial_operating_nol"), "$mm", "Parallel tax schedules"],
        ["Risk-free rate", a.get("risk_free_rate"), "%", "Currency matched"],
        ["Market risk premium", a.get("market_risk_premium"), "%", "Market premium"],
        ["Selected asset beta", a.get("selected_asset_beta"), "beta", "Comparable-company operating risk"],
        ["Unlevered TOCC", s.get("tocc"), "%", "Risk-free + adjusted asset beta × MRP"],
        ["Terminal growth", a.get("terminal_growth_rate"), "%", "Must be below TOCC"],
        ["Terminal RONIC", s.get("terminal_ronic"), "%", "Converges to TOCC"],
        ["Minimum cash", a.get("minimum_cash"), "$mm", "Liquidity threshold"],
        ["Interest limitation", a.get("interest_limit_percentage"), "% ATI", "Interest deductibility convention"],
        ["Annual SBC dilution", a.get("annual_sbc_dilution_rate"), "%", "Share-count schedule"],
    ]
    first, last = _table(ws, ["Assumption", "Value", "Units", "Notes"], assumption_rows)
    for row in range(first, last + 1):
        _input(ws.cell(row, 2))
    for row in [7, 8, 9, 10, 11, 13, 14, 16, 17, 20, 21]:
        ws.cell(row, 2).number_format = PERCENT
    _set_widths(ws, {1: 34, 2: 20, 3: 16, 4: 52})

    _timeseries(workbook["Revenue Build"], "Revenue Build", "Explicit and competitive-advantage revenue forecast", [*f, *op], [("Revenue growth", "revenue_growth", PERCENT), ("Revenue", "revenue", MONEY), ("Revenue change", "net_investment", MONEY)], fill=FORECAST)
    _timeseries(workbook["PP&E Schedule"], "PP&E Schedule", "Fixed asset requirement and capital-expenditure roll-forward", f, [("Revenue", "revenue", MONEY), ("Net PP&E", "net_ppe", MONEY), ("D&A", "da", MONEY), ("Capital expenditures", "capex", MONEY)], fill=FORECAST)
    _timeseries(workbook["Cost Build"], "Cost Build", "Operating-cost and margin assumptions", f, [("Revenue", "revenue", MONEY), ("COGS", "cogs", MONEY), ("Gross profit", "gross_profit", MONEY), ("SG&A", "sga", MONEY), ("R&D", "rd", MONEY), ("D&A", "da", MONEY), ("EBIT margin", "ebit_margin", PERCENT), ("EBIT", "ebit", MONEY)], fill=FORECAST)
    _timeseries(workbook["IS Forecast"], "Income Statement Forecast", "Driver-based operating income forecast", f, [("Revenue", "revenue", MONEY), ("COGS", "cogs", MONEY), ("Gross profit", "gross_profit", MONEY), ("SG&A", "sga", MONEY), ("R&D", "rd", MONEY), ("D&A", "da", MONEY), ("EBIT", "ebit", MONEY), ("Operating taxes", "operating_taxes", MONEY), ("NOPAT", "nopat", MONEY)], fill=FORECAST)
    _timeseries(workbook["BS Forecast"], "Balance Sheet Forecast", "Operating assets, working capital, liquidity, and debt", f, [("Operating working capital", "owc", MONEY), ("Net PP&E", "net_ppe", MONEY), ("Operating invested capital", "invested_capital", MONEY), ("Opening cash", "opening_cash", MONEY), ("Ending cash", "ending_cash", MONEY), ("Opening debt", "opening_debt", MONEY), ("Ending debt", "ending_debt", MONEY)], fill=FORECAST)

    # DCF and cash-flow schedules
    _timeseries(workbook["CF Forecast"], "Cash Flow Forecast", "Unlevered free cash flow and present value", f, [("NOPAT", "nopat", MONEY), ("D&A", "da", MONEY), ("Capital expenditures", "capex", MONEY), ("Change in operating working capital", "change_owc", MONEY), ("Net new investment", "net_investment", MONEY), ("Unlevered FCF", "fcf", MONEY), ("Discount factor", "discount_factor", "0.000x"), ("PV of FCF", "pv_fcf", MONEY)], fill=FORECAST)

    ws = workbook["DCF"]
    _title(ws, "Discounted Cash Flow", "Operating value before separately valued financing effects", max(8, len(f) + 1))
    _table(ws, ["Metric", *[row.get("year") for row in f]], [
        ["Unlevered FCF", *[row.get("fcf") for row in f]],
        ["Discount factor", *[row.get("discount_factor") for row in f]],
        ["PV of explicit FCF", *[row.get("pv_fcf") for row in f]],
    ])
    end_col = len(f) + 1
    _format(ws, 5, 5, 2, end_col, MONEY)
    _format(ws, 6, 6, 2, end_col, "0.000x")
    _format(ws, 7, 7, 2, end_col, MONEY)
    summary_row = 10
    dcf_rows = [
        ["PV of explicit forecast FCF", s.get("pv_explicit_fcf")],
        ["PV of competitive-advantage FCF", s.get("pv_overperformance_fcf")],
        ["PV of terminal value", s.get("pv_terminal_value")],
        ["PV of continuing value", "=SUM(B11:B12)"],
        ["Operating enterprise value", "=SUM(B10,B13)"],
    ]
    _table(ws, ["DCF summary", "Value"], dcf_rows, start_row=summary_row, freeze=None)
    _format(ws, summary_row + 1, summary_row + 5, 2, 2, MONEY)
    _set_widths(ws, {1: 38, **{column: 14 for column in range(2, end_col + 1)}})

    _timeseries(workbook["ROIC Tree - Forecast"], "ROIC Tree - Forecast", "Forecast returns and value creation", [*f, *op], [("Revenue growth", "revenue_growth", PERCENT), ("EBIT margin", "ebit_margin", PERCENT), ("NOPAT", "nopat", MONEY), ("Capital turnover", "capital_turnover", MULTIPLE), ("ROIC", "roic", PERCENT), ("RONIC", "ronic", PERCENT), ("Economic profit", "economic_profit", MONEY), ("Unlevered FCF", "fcf", MONEY)], fill=FORECAST)

    # APV, tax shields, and financing
    ws = workbook["APV Method"]
    _title(ws, "Adjusted Present Value Method", "Operating value plus separately valued financing effects", 6)
    apv_rows = [
        ["Unlevered TOCC", s.get("tocc")],
        ["Terminal growth", s.get("terminal_growth")],
        ["Terminal RONIC", s.get("terminal_ronic")],
        ["PV explicit FCF", s.get("pv_explicit_fcf")],
        ["PV continuing value", s.get("pv_continuing_value")],
        ["Operating enterprise value", "=SUM(B8:B9)"],
        ["PV explicit interest tax shields", s.get("pv_explicit_tax_shields")],
        ["PV continuing interest tax shield", s.get("pv_continuing_tax_shield")],
        ["PV other financing effects", s.get("pv_other_financing_effects")],
        ["PV financing effects", "=SUM(B11:B13)"],
        ["APV enterprise value", "=B10+B14"],
        ["PV continuing value / APV EV", '=IF(B15=0,"",B9/B15)'],
        ["Less: gross debt", s.get("gross_debt")],
        ["Less: other financing claims", s.get("other_financing_claims")],
        ["Add: excess cash and investments", s.get("excess_cash")],
        ["Other non-operating adjustments", company.get("non_operating_investments", 0)],
        ["Equity value", "=MAX(0,B15-B17-B18+B19+B20)"],
        ["Diluted shares", s.get("diluted_shares")],
        ["Intrinsic value / share", '=IF(B22=0,"",B21/B22)'],
        ["Market price", s.get("market_price")],
        ["Market capitalization", s.get("market_cap")],
        ["Premium / (discount)", '=IF(B24=0,"",B23/B24-1)'],
    ]
    first, last = _table(ws, ["APV and equity bridge", "Value"], apv_rows)
    _format(ws, 5, 7, 2, 2, PERCENT)
    _format(ws, 8, 21, 2, 2, MONEY)
    _format(ws, 22, 22, 2, 2, COUNT)
    _format(ws, 23, 24, 2, 2, PER_SHARE)
    _format(ws, 25, 25, 2, 2, MONEY)
    _format(ws, 26, 26, 2, 2, PERCENT)
    for row in [10, 14, 15, 16, 21, 23, 26]:
        ws.cell(row, 2).font = Font(bold=True, color=LINK_GREEN if row in [10, 14, 15, 21, 23, 26] else "000000")
    _set_widths(ws, {1: 48, 2: 24})

    ws = workbook["APV Tax Shields"]
    _title(ws, "APV Tax Shields", "Interest deductibility, NOLs, and incremental financing benefit", 18)
    headers = ["Year", "Phase", "Interest", "ATI", "ATI limit", "Opening interest CF", "CF used", "Deductible interest", "Ending interest CF", "Opening NOL w/o interest", "NOL used w/o interest", "Ending NOL w/o interest", "Tax w/o interest", "Opening NOL with interest", "NOL used with interest", "Ending NOL with interest", "Tax with interest", "Usable tax shield"]
    shield_rows = [[row.get("year"), row.get("phase"), row.get("interest"), row.get("ati"), row.get("limit"), row.get("opening_carryforward"), row.get("carryforward_used"), row.get("deductible_interest"), row.get("ending_carryforward"), row.get("opening_nol_without_interest"), row.get("nol_used_without_interest"), row.get("ending_nol_without_interest"), row.get("cash_tax_without_interest"), row.get("opening_nol_with_interest"), row.get("nol_used_with_interest"), row.get("ending_nol_with_interest"), row.get("cash_tax_with_interest"), row.get("usable_tax_shield")] for row in result.tax_shield]
    first, last = _table(ws, headers, shield_rows)
    _format(ws, first, last, 3, 18, MONEY)
    _set_widths(ws, {1: 12, 2: 20, **{column: 18 for column in range(3, 19)}})

    ws = workbook["Financing Plan"]
    _title(ws, "Financing Plan and Liquidity", "Cash and debt roll-forward; funding needs are disclosed, not plugged", 10)
    periods = [*f, *op]
    financing_rows = [[row.get("year"), row.get("phase"), row.get("opening_cash"), row.get("fcf"), row.get("new_borrowing"), row.get("equity_raise"), row.get("mandatory_repayment"), row.get("interest_expense"), row.get("ending_debt"), row.get("ending_cash")] for row in periods]
    first, last = _table(ws, ["Year", "Phase", "Opening cash", "UFCF", "New borrowing", "Equity raise", "Debt repayment", "Cash interest", "Ending debt", "Ending cash"], financing_rows)
    _format(ws, first, last, 3, 10, MONEY)
    summary_start = last + 3
    liquidity_rows = [
        ["Minimum modeled cash", s.get("minimum_cash_balance")],
        ["Required minimum cash", a.get("minimum_cash")],
        ["Minimum external funding", s.get("minimum_external_funding")],
        ["First breach year", s.get("first_liquidity_breach_year")],
        ["Years below minimum", s.get("years_below_minimum_cash")],
    ]
    _table(ws, ["Liquidity check", "Value"], liquidity_rows, start_row=summary_start, freeze=None)
    _format(ws, summary_start + 1, summary_start + 3, 2, 2, MONEY)
    _set_widths(ws, {1: 28, 2: 20, **{column: 18 for column in range(3, 11)}})

    # Scenarios and model checks
    ws = workbook["Scenarios"]
    _title(ws, "Valuation Scenarios", "Only probability cells are intended for user editing", 10)
    scenario_rows = []
    for row in result.scenarios:
        scenario_rows.append([row.get("scenario"), row.get("probability"), "Liquidation" if row.get("liquidation") else "Going concern", row.get("tocc"), row.get("terminal_growth"), row.get("apv_enterprise_value"), row.get("equity_value"), row.get("diluted_shares"), row.get("intrinsic_value_per_share"), None])
    first, last = _table(ws, ["Scenario", "Probability", "Type", "TOCC", "Terminal growth", "APV EV", "Equity value", "Diluted shares", "Intrinsic value / share", "Probability-weighted value"], scenario_rows)
    for row in range(first, last + 1):
        _input(ws.cell(row, 2))
        ws.cell(row, 10, f"=B{row}*I{row}")
    _format(ws, first, last, 2, 2, PERCENT)
    _format(ws, first, last, 4, 5, PERCENT)
    _format(ws, first, last, 6, 7, MONEY)
    _format(ws, first, last, 8, 8, COUNT)
    _format(ws, first, last, 9, 10, PER_SHARE)
    total_row = last + 2
    ws.cell(total_row, 1, "Probability total")
    ws.cell(total_row, 2, f"=SUM(B{first}:B{last})")
    ws.cell(total_row, 8, "Expected intrinsic value")
    ws.cell(total_row, 9, f"=SUM(J{first}:J{last})")
    ws.cell(total_row, 2).number_format = PERCENT
    ws.cell(total_row, 9).number_format = PER_SHARE
    ws.conditional_formatting.add(f"B{total_row}", CellIsRule(operator="notEqual", formula=[1], fill=PatternFill("solid", fgColor=RED)))
    _set_widths(ws, {column: 19 for column in range(1, 11)})

    ws = workbook["Model Checks"]
    _title(ws, "Model Checks", "Every check returns PASS, WARNING, or FAIL with an explanatory note", 8)
    ws["A4"] = "Overall model status"
    ws["B4"] = s.get("overall_model_status")
    for cell in ws[4][0:2]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(color=WHITE, bold=True)
    check_rows = [[row.category, row.name, row.actual, row.expected, row.difference, row.tolerance, row.status, row.notes] for row in result.checks]
    first, last = _table(ws, ["Category", "Check", "Actual", "Expected", "Difference", "Tolerance", "Status", "Notes"], check_rows, start_row=6)
    _status_rules(ws, "B4")
    if last >= first:
        _status_rules(ws, f"G{first}:G{last}")
    _set_widths(ws, {1: 22, 2: 48, 3: 18, 4: 18, 5: 18, 6: 18, 7: 16, 8: 54})

    ws = workbook["Sources"]
    _title(ws, "Sources and Provenance", "Imported values and assumptions remain traceable", 10)
    source_rows = [[row.get("variable"), row.get("value"), row.get("source"), row.get("source_date"), row.get("retrieval_method"), row.get("original_unit"), row.get("normalized_unit"), "Yes" if row.get("user_override") else "No", row.get("confidence"), row.get("notes")] for row in result.provenance]
    first, last = _table(ws, ["Variable", "Value", "Source", "Source date", "Retrieval method", "Original unit", "Normalized unit", "User override?", "Confidence", "Notes"], source_rows)
    for row in ws.iter_rows(min_row=first, max_row=last, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    _set_widths(ws, {1: 30, 2: 20, 3: 62, 4: 18, 5: 24, 6: 18, 7: 18, 8: 18, 9: 16, 10: 60})

    # Add concise, source-linked charts to the valuation page.
    chart_data_start = 22
    ws = workbook["Valuation"]
    combined = [
        {"year": row.get("year"), "revenue": row.get("revenue"), "ebit_margin": row.get("ebit_margin"), "roic": row.get("roic"), "tocc": row.get("tocc"), "fcf": row.get("fcf")}
        for row in h
    ] + [
        {"year": row.get("year"), "revenue": row.get("revenue"), "ebit_margin": row.get("ebit_margin"), "roic": row.get("roic"), "tocc": s.get("tocc"), "fcf": row.get("fcf")}
        for row in [*f, *op]
    ]
    headers = ["Year", "Revenue", "EBIT margin", "ROIC", "TOCC", "UFCF"]
    for col, header in enumerate(headers, start=1):
        ws.cell(chart_data_start, col, header)
    _header(ws, chart_data_start, 1, len(headers))
    for row_index, row in enumerate(combined, start=chart_data_start + 1):
        for col, key in enumerate(["year", "revenue", "ebit_margin", "roic", "tocc", "fcf"], start=1):
            ws.cell(row_index, col, row.get(key))
    chart_last = chart_data_start + len(combined)
    if combined:
        revenue_chart = LineChart()
        revenue_chart.title = "Revenue ($mm)"
        revenue_chart.style = 13
        revenue_chart.height = 7
        revenue_chart.width = 12
        revenue_chart.add_data(Reference(ws, min_col=2, min_row=chart_data_start, max_row=chart_last), titles_from_data=True)
        revenue_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_last))
        revenue_chart.y_axis.numFmt = '$#,##0'
        ws.add_chart(revenue_chart, "H22")

        roic_chart = LineChart()
        roic_chart.title = "ROIC versus TOCC"
        roic_chart.style = 13
        roic_chart.height = 7
        roic_chart.width = 12
        roic_chart.add_data(Reference(ws, min_col=4, max_col=5, min_row=chart_data_start, max_row=chart_last), titles_from_data=True)
        roic_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_last))
        roic_chart.y_axis.numFmt = '0.0%'
        ws.add_chart(roic_chart, "H37")

        fcf_chart = BarChart()
        fcf_chart.title = "Unlevered FCF ($mm)"
        fcf_chart.style = 10
        fcf_chart.height = 7
        fcf_chart.width = 12
        fcf_chart.add_data(Reference(ws, min_col=6, min_row=chart_data_start, max_row=chart_last), titles_from_data=True)
        fcf_chart.set_categories(Reference(ws, min_col=1, min_row=chart_data_start + 1, max_row=chart_last))
        fcf_chart.y_axis.numFmt = '$#,##0'
        ws.add_chart(fcf_chart, "H52")

    _format(ws, chart_data_start + 1, chart_last, 2, 2, MONEY)
    _format(ws, chart_data_start + 1, chart_last, 3, 5, PERCENT)
    _format(ws, chart_data_start + 1, chart_last, 6, 6, MONEY)

    # Workbook-level finishing and navigation.
    for ws in workbook.worksheets:
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_view.zoomScale = 90
        ws.auto_filter.ref = None
    workbook["Valuation"].sheet_view.zoomScale = 85
    workbook.active = workbook.sheetnames.index("Valuation")
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(path)
    return path
