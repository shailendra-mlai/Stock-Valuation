from openpyxl import load_workbook

from valuation_system.analysis.engine import run_valuation
from valuation_system.data.company_data import load_company_data
from valuation_system.models.assumptions import ValuationAssumptions
from valuation_system.reporting.cloud_excel import TABS, export_cloud_excel


def test_cloud_excel_matches_reference_style_architecture(tmp_path):
    company = load_company_data("DEMO", "sample_company_data.csv", False, "test")
    company.comparables = []
    result = run_valuation(company, ValuationAssumptions())

    output = export_cloud_excel(result, tmp_path / "DEMO_Valuation.xlsx")
    workbook = load_workbook(output, data_only=False)

    assert workbook.sheetnames == TABS
    assert workbook["Valuation"]["B5"].value == "='APV Method'!B5"
    assert workbook["Valuation"]["B18"].value == "='Model Checks'!B4"
    assert workbook["APV Method"]["B15"].value == "=B10+B14"
    assert workbook["Scenarios"]["B5"].value == 0.25
    assert workbook["Scenarios"]["J5"].value == "=B5*I5"
    assert len(workbook["Valuation"]._charts) == 3
    assert workbook.calculation.fullCalcOnLoad is True

